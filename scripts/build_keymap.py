"""매칭표 엑셀을 검색이 읽는 사전으로 바꾼다.

## 왜 변환이 필요한가

사람이 만든 매칭표는 엑셀이고 사람이 보기 좋게 되어 있다. 검색은 그걸 못
읽는다. 낱말 하나에 절 하나를 잇는 평평한 표로 펼쳐야 한다.

    매칭표 한 행                     이 사전의 여러 행
    열쇠뜻 "배당에 관한 사항"          결산배당 → 배당에관한사항
    AI남길것 "결산배당, 배당률, …"     배당률   → 배당에관한사항
                                     …

## 경로를 쓰지 않는 이유

`pathmap.csv` 는 경로(`III/6`)를 열쇠로 쓴다. 그런데 경로는 회사마다 다른
절을 가리킨다. 실측했다.

    III/6 에 들어앉은 절   18가지
       배당에 관한 사항(70곳) · 범주별 금융상품(33곳) · 매출채권(21곳) …

    경로당 절 수 (평균)
       I IV V VI VIII IX X XI   1.0   안정적이다
       III                      5.0   최대 25.  쓸 수 없다

III 가 절의 대부분인데 하필 거기가 경로를 못 믿는 자리다. 그래서 경로 대신
절의 정체를 열쇠로 쓴다. `build_section_schema.py` 의 `schema_key` 와 같은
규칙이다.

    III 이고 쓸 만한 XBRL 태그가 있으면   그 태그
    그 밖                                제목에서 앞 번호를 뗀 것

태그를 쓰는 이유는 회사가 태그를 안 바꾸기 때문이다. 같은 종속기업투자를
31가지 이름으로 부르는데 태그는 하나다.

## 낱말은 어디서 가져오나

`AI남길것` 칸이다. 기계가 뽑은 후보를 사람이 손본 결과가 여기 들어 있다.
2026-09-04 판에서 사람이 67종을 새로 넣고 89종을 뺐다. 새로 넣은 것이
감사위원회·우발부채·배당률·주주총회의사록처럼 사람이 실제로 쓰는 말이다.

`낱말` 칸(기계가 뽑은 원본)은 안 쓴다. 그쪽은 표의 머리글이 대부분이라
질의에 안 나온다.

## 만드는 것

    data/eval/keymap.csv
        word · major · key_type · key · key_label · n_corp · n_key

    n_key 는 그 낱말이 가리키는 절의 수다. 1 이면 확실하고, 여럿이면
    그 절들을 다 올린다. 대부분 같은 계열의 형제 절이다.

## 쓰는 법

    python scripts/build_keymap.py
    python scripts/build_keymap.py --xlsx=reference/매칭표_최종본.xlsx
"""
from __future__ import annotations

import csv
import glob
import sys
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))
sys.stdout.reconfigure(encoding="utf-8")

OUT = ROOT / "data" / "eval" / "keymap.csv"

# 낱말로 쓰기에 부적절한 것. 사람이 지우고 남은 잡음을 한 번 더 거른다.
#
# 신용등급 기호(CCC·BB·AAA)와 통화 기호(USD·EUR)가 특히 성가시다. 낱말처럼
# 생겨서 눈으로 훑으면 지나치는데, 이런 기호가 든 절이 몇 개 없어 점수가
# 높게 잡힌다. 질의에 우연히 영문 세 글자가 들어가면 엉뚱한 절이 올라온다.
BAD_EXACT = {
    "CCC", "CC", "C", "BB", "B", "BBB", "AAA", "AA", "A", "D",
    "USD", "EUR", "JPY", "CNY", "KRW", "and", "or", "the",
    "LLC", "LTD", "SPC", "SPV", "PLC", "INC", "CO",
}
MIN_LEN = 2         # 한 글자 낱말은 안 쓴다


# 매칭표에 빠져 있어 손으로 채운 것.
#
# 2026-09-05 안전성 시험에서 "등기임원 전원의 이름과 생년월일" 질의가
# 아무 절도 못 짚었다. 매칭표를 뒤져 보니 임원·등기임원·임직원·이사·
# 직위·약력이 하나도 없었다. 사람이 표를 만들 때 III 재무 주석에 집중해
# VIII 임원 쪽이 비었다.
#
# 열쇠는 실제 절 이름을 정규화한 값이다. `build_section_schema.py` 의
# `norm()` 과 같아야 검색이 조각에서 되찾을 수 있다. 확인한 절이다.
#
#     VI/1     1. 이사회에 관한 사항        70개사
#     VIII/1   1. 임원 및 직원 등의 현황     70개사
#     VIII/2   2. 임원의 보수 등           70개사
#
# 표를 직접 고치지 않고 여기서 더한다. 매칭표는 사람이 만드는 것이라
# 기계가 끼어들면 다음에 사람이 고칠 때 무엇이 자기 것인지 헷갈린다.
# 여기 두면 출처가 분명하고 되돌리기도 쉽다.
EXTRA: list[tuple[str, str, str, str, int]] = [
    # (낱말, 대분류, 열쇠종류, 열쇠, 기업수)
    ("임원", "VIII", "title", "임원및직원등의현황", 70),
    ("등기임원", "VIII", "title", "임원및직원등의현황", 70),
    ("임직원", "VIII", "title", "임원및직원등의현황", 70),
    ("직원현황", "VIII", "title", "임원및직원등의현황", 70),
    ("근속연수", "VIII", "title", "임원및직원등의현황", 70),
    ("출생연월", "VIII", "title", "임원및직원등의현황", 70),
    ("약력", "VIII", "title", "임원및직원등의현황", 70),
    ("겸직", "VIII", "title", "임원및직원등의현황", 70),
    ("임원보수", "VIII", "title", "임원의보수등", 70),
    ("보수총액", "VIII", "title", "임원의보수등", 70),
    ("연봉", "VIII", "title", "임원의보수등", 70),
    ("보수한도", "VIII", "title", "임원의보수등", 70),
    ("이사회", "VI", "title", "이사회에관한사항", 70),
    ("사외이사", "VI", "title", "이사회에관한사항", 70),
    ("이사선임", "VI", "title", "이사회에관한사항", 70),
    ("최대주주", "VII", "title", "주주에관한사항", 70),
    ("주주현황", "VII", "title", "주주에관한사항", 70),
    ("지분율", "VII", "title", "주주에관한사항", 70),
]


def bad(w: str) -> bool:
    """낱말로 못 쓰는 것인가."""
    if len(w) < MIN_LEN or w in BAD_EXACT:
        return True
    if w.replace(",", "").replace(".", "").isdigit():
        return True
    return False


def cells(cell) -> list[str]:
    """쉼표로 이어진 칸을 낱말 목록으로."""
    return [x.strip() for x in str(cell or "").split(",") if x.strip()]


def find_xlsx() -> Path:
    """매칭표 파일. 최종본을 우선한다."""
    got = sorted(glob.glob(str(ROOT / "reference" / "*.xlsx")))
    fin = [g for g in got if "최종" in g]
    if fin:
        return Path(fin[0])
    if not got:
        raise SystemExit("reference/ 에 xlsx 가 없다")
    return Path(got[-1])


def main(xlsx: Path) -> int:
    import openpyxl

    ws = openpyxl.load_workbook(xlsx, data_only=True).active
    head = [c.value for c in ws[1]]
    ix = {h: i for i, h in enumerate(head)}
    need = ["대분류", "열쇠", "열쇠뜻", "열쇠종류", "기업수", "AI남길것"]
    miss = [n for n in need if n not in ix]
    if miss:
        raise SystemExit(f"칸이 없다: {miss}")

    rows = [[c.value for c in r] for r in ws.iter_rows(min_row=2)]
    print(f"매칭표 {xlsx.name} · {len(rows)}행")

    # 낱말 → 그 낱말이 가리키는 절들
    byword: dict[str, list[tuple] ] = defaultdict(list)
    n_word = n_bad = n_row = 0
    for r in rows:
        key = r[ix["열쇠"]]
        if not key:
            continue
        unit = (str(r[ix["대분류"]]), str(r[ix["열쇠종류"]]), str(key),
                str(r[ix["열쇠뜻"]]), int(r[ix["기업수"]] or 0))
        ws_ = cells(r[ix["AI남길것"]])
        if not ws_:
            continue
        n_row += 1
        for w in ws_:
            n_word += 1
            if bad(w):
                n_bad += 1
                continue
            byword[w].append(unit)

    print(f"낱말 자리 {n_word:,} · 잡음으로 뺀 것 {n_bad} · 남은 낱말 {len(byword):,}종")
    print(f"낱말이 붙은 행 {n_row} / {len(rows)}")

    n_new = 0
    for w, top, kt, key, nc in EXTRA:
        unit = (top, kt, key, key, nc)
        if unit not in byword[w]:
            byword[w].append(unit)
            n_new += 1
    print(f"손으로 채운 것 {n_new}줄 (임원·이사회·주주. EXTRA 참조)")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with OUT.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["word", "major", "key_type", "key", "key_label",
                    "n_corp", "n_key"])
        for word in sorted(byword):
            units = byword[word]
            for u in units:
                w.writerow([word, u[0], u[1], u[2], u[3], u[4], len(units)])

    one = sum(1 for v in byword.values() if len(v) == 1)
    print(f"\n{OUT.relative_to(ROOT)} 를 만들었다")
    print(f"   절 하나만 가리키는 낱말  {one:,}종  ({one / len(byword) * 100:.0f}%)")
    print(f"   여럿을 가리키는 낱말     {len(byword) - one:,}종")
    print(f"   전체 행                {sum(len(v) for v in byword.values()):,}")
    return 0


if __name__ == "__main__":
    p = None
    for a in sys.argv[1:]:
        if a.startswith("--xlsx="):
            p = Path(a.split("=", 1)[1])
    sys.exit(main(p or find_xlsx()))
